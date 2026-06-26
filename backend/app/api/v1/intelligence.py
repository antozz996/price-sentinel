"""
Price Sentinel — Intelligence Router (Sprint 4).
Espone gli endpoint per la Dashboard Admin, KPI e Cross-Location Tracker.
"""

from datetime import date, timedelta, datetime, timezone
from typing import Any
from io import BytesIO

from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy import select, func, and_, or_, case, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import require_admin, get_current_user
from app.database import get_db
from app.models.anomalie import Anomalia, NotaDiCredito, StatoValidazione, ApprovazionePrezzo
from app.models.fatture import RigaFattura, Fattura, StatoMatching
from app.models.location import Location
from app.models.listino import ListinoMaster
from app.models.fornitori import Fornitore
from fastapi.responses import Response, StreamingResponse
from app.services.pdf_generator import generate_vendor_passport_pdf, generate_consumption_invoices_pdf
from app.schemas.approvazioni import ApprovazionePrezzoCreate, ApprovazionePrezzoResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()


@router.get("/kpi", summary="Kpi Economici Principali")
async def get_kpi(
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_admin),
) -> dict[str, Any]:
    """
    Ritorna gli indicatori economici:
    - Euro Recuperati Totali (NC registrate)
    - Euro In Contestazione (Anomalie in_reclamo)
    - Euro A Rischio (Anomalie contestate ma non ancora in reclamo)
    """
    # Optimized single query for Anomalia states
    anomalie_stmt = select(
        func.coalesce(func.sum(case((Anomalia.stato_validazione == StatoValidazione.in_reclamo, Anomalia.delta_totale), else_=0)), 0).label("in_contestazione"),
        func.coalesce(func.sum(case((Anomalia.stato_validazione == StatoValidazione.contestata, Anomalia.delta_totale), else_=0)), 0).label("a_rischio"),
        func.coalesce(func.sum(case((Anomalia.stato_validazione == StatoValidazione.da_verificare, Anomalia.delta_totale), else_=0)), 0).label("attesa_manager"),
    )
    anomalie_res = await db.execute(anomalie_stmt)
    anomalie_row = anomalie_res.one()

    # Recuperati Totali (from NotaDiCredito table)
    recuperati_res = await db.execute(
        select(func.coalesce(func.sum(NotaDiCredito.importo_recuperato), 0))
    )
    recuperati = recuperati_res.scalar()

    return {
        "euro_recuperati": float(recuperati),
        "euro_in_contestazione": float(anomalie_row.in_contestazione),
        "euro_a_rischio": float(anomalie_row.a_rischio),
        "euro_attesa_manager": float(anomalie_row.attesa_manager)
    }


@router.get("/expiring-protocols", summary="Widget Scadenze Contratti")
async def get_expiring_protocols(
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_admin),
):
    """
    Lista semaforica dei ListinoMaster in scadenza nei prossimi 30 giorni.
    """
    oggi = date.today()
    scadenza_limite = oggi + timedelta(days=30)
    
    res = await db.execute(
        select(ListinoMaster)
        .where(and_(
            ListinoMaster.data_scadenza <= scadenza_limite,
            ListinoMaster.data_scadenza >= oggi
        ))
        .order_by(ListinoMaster.data_scadenza.asc())
    )
    listini = res.scalars().all()
    
    result = []
    for l in listini:
        days_left = (l.data_scadenza - oggi).days
        color = "red" if days_left < 7 else ("yellow" if days_left < 15 else "green")
        result.append({
            "id": l.id,
            "sku_interno": l.sku_interno,
            "fornitore_id": l.fornitore_id,
            "data_scadenza": l.data_scadenza,
            "days_left": days_left,
            "color": color
        })
    return result


@router.get("/cross-location", summary="Cross-Location Tracker")
async def get_cross_location_matrix(
    data_da: date | None = Query(None, description="Data inizio"),
    data_a: date | None = Query(None, description="Data fine"),
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_admin),
):
    """
    Matrice di intelligence negoziale.
    Ritorna l'ultimo prezzo di acquisto di uno sku per ogni location.
    I dati permettono al frontend di costruire una griglia Prodotti x Location.
    """
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None

    # Costruzione condizioni dinamiche per data
    conditions = [
        RigaFattura.stato_matching == StatoMatching.matched,
        RigaFattura.sku_interno.isnot(None),
        RigaFattura.prezzo_netto_normalizzato > 0,
        RigaFattura.is_omaggio.isnot(True)
    ]
    if data_da:
        conditions.append(Fattura.data_documento >= data_da)
    if data_a:
        conditions.append(Fattura.data_documento <= data_a)

    # Vogliamo l'ultimo prezzo normalizzato per (sku_interno, location_id).
    # Group By SKU e Location, max ID
    subquery = (
        select(
            RigaFattura.sku_interno,
            Fattura.location_id,
            func.max(RigaFattura.id).label("max_riga_id")
        )
        .join(Fattura, RigaFattura.fattura_id == Fattura.id)
        .where(and_(*conditions))
        .group_by(RigaFattura.sku_interno, Fattura.location_id)
        .subquery()
    )
    
    stmt = (
        select(
            RigaFattura.sku_interno, 
            Fattura.location_id, 
            RigaFattura.prezzo_netto_normalizzato,
            RigaFattura.descrizione_fornitore_raw,
            Fattura.id.label("fattura_id"),
            RigaFattura.quantita
        )
        .join(subquery, RigaFattura.id == subquery.c.max_riga_id)
        .join(Fattura, RigaFattura.fattura_id == Fattura.id)
    )
    
    res = await db.execute(stmt)
    records = res.all()
    
    # Costruiamo la risposta JSON formattata per la griglia UI
    matrix = {}
    for sku, loc_id, price, desc, fattura_id, quantita in records:
        display_name = f"{desc or 'Prodotto Senza Nome'} ({sku})"
        if display_name not in matrix:
            matrix[display_name] = {}
        # Arrotonda a 2 decimali per allinearsi perfettamente alla griglia UI ed evitare falsi positivi
        matrix[display_name][loc_id] = {
            "prezzo": round(float(price), 2),
            "fattura_id": int(fattura_id),
            "quantita": float(quantita)
        }
        
    # Filtriamo per restituire solo gli SKU con reale delta prezzi tra le sedi (delta > 0.01)
    filtered_matrix = {}
    for display_name, loc_prices in matrix.items():
        prices = [item["prezzo"] for item in loc_prices.values()]
        if len(prices) > 1 and (max(prices) - min(prices)) > 0.011:
            filtered_matrix[display_name] = loc_prices
            
    # Ordiniamo alfabeticamente per nome del prodotto
    sorted_matrix = {k: filtered_matrix[k] for k in sorted(filtered_matrix.keys())}
            
    return sorted_matrix


@router.get("/cross-supplier", summary="Cross-Supplier Pricing Matrix")
async def get_cross_supplier_matrix(
    data_da: date | None = Query(None, description="Data inizio"),
    data_a: date | None = Query(None, description="Data fine"),
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_admin),
):
    """
    Ritorna la matrice comparativa incrociata dei prezzi per fornitore.
    Incrocia i listini concordati (ListinoMaster) e le fatture storiche (prezzo spot minimo).
    """
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None

    # 1. Recupero contratti attivi nel periodo
    contracts_stmt = select(
        ListinoMaster.sku_interno,
        ListinoMaster.descrizione,
        ListinoMaster.fornitore_id,
        ListinoMaster.prezzo_pattuito
    )
    contracts_conds = []
    if data_da:
        contracts_conds.append(or_(ListinoMaster.data_scadenza.is_(None), ListinoMaster.data_scadenza >= data_da))
    if data_a:
        contracts_conds.append(ListinoMaster.data_inizio_validita <= data_a)
        
    if not contracts_conds:
        contracts_conds.append(ListinoMaster.data_scadenza.is_(None))
        
    contracts_stmt = contracts_stmt.where(and_(*contracts_conds))
    
    contracts_res = await db.execute(contracts_stmt)
    contracts = contracts_res.all()
    
    # 2. Recupero prezzi storici spot minimi da righe fattura registrate (matched)
    # Escludiamo omaggi e articoli a prezzo zero (prezzo_netto_normalizzato > 0 e is_omaggio is False)
    spot_conds = [
        RigaFattura.stato_matching == StatoMatching.matched,
        RigaFattura.sku_interno.isnot(None),
        RigaFattura.prezzo_netto_normalizzato > 0,
        RigaFattura.is_omaggio.isnot(True)
    ]
    if data_da:
        spot_conds.append(Fattura.data_documento >= data_da)
    if data_a:
        spot_conds.append(Fattura.data_documento <= data_a)

    spot_stmt = (
        select(
            RigaFattura.sku_interno,
            RigaFattura.descrizione_fornitore_raw,
            Fattura.fornitore_id,
            func.min(RigaFattura.prezzo_netto_normalizzato).label("prezzo_min")
        )
        .join(Fattura, RigaFattura.fattura_id == Fattura.id)
        .where(and_(*spot_conds))
        .group_by(RigaFattura.sku_interno, RigaFattura.descrizione_fornitore_raw, Fattura.fornitore_id)
    )
    
    spot_res = await db.execute(spot_stmt)
    spots = spot_res.all()
    
    # 3. Consolidamento dei dati in formato SKU -> prezzi per fornitore
    matrix = {}
    
    # Processiamo prima gli spot storici
    for sku, desc, fornitore_id, prezzo_min in spots:
        if not sku:
            continue
        if sku not in matrix:
            matrix[sku] = {
                "sku_interno": sku,
                "descrizione": desc or f"Prodotto {sku}",
                "prezzi": {}
            }
        matrix[sku]["prezzi"][str(fornitore_id)] = {
            "prezzo": round(float(prezzo_min), 2),
            "tipo": "spot"
        }
        
    # Sovrapponiamo i contratti concordati (hanno priorità rispetto allo spot dello stesso fornitore)
    for sku, desc, fornitore_id, prezzo_pattuito in contracts:
        if not sku:
            continue
        if sku not in matrix:
            matrix[sku] = {
                "sku_interno": sku,
                "descrizione": desc or f"Prodotto {sku}",
                "prezzi": {}
            }
        if desc:
            matrix[sku]["descrizione"] = desc
            
        matrix[sku]["prezzi"][str(fornitore_id)] = {
            "prezzo": round(float(prezzo_pattuito), 2),
            "tipo": "concordato"
        }
        
    # Ordiniamo alfabeticamente per descrizione prodotto
    sorted_matrix = {}
    for sku in sorted(matrix.keys(), key=lambda s: matrix[s]["descrizione"].lower()):
        sorted_matrix[sku] = matrix[sku]
        
    return sorted_matrix


@router.get("/export-vendor-passport/{fornitore_id}", summary="Download Vendor Passport PDF")
async def export_vendor_passport(
    fornitore_id: int,
    data_da: date | None = Query(None, description="Data inizio"),
    data_a: date | None = Query(None, description="Data fine"),
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_admin),
):
    """
    Genera il PDF Business Intelligence per il fornitore.
    """
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None
    from app.models.fornitori import Fornitore
    fornitore = await db.scalar(select(Fornitore).where(Fornitore.id == fornitore_id))
    if not fornitore:
        from fastapi import HTTPException
        raise HTTPException(404, "Fornitore non trovato")
        
    # Calcolo assortimento aggregato senza esporre prezzi
    stmt = (
        select(Location.nome_struttura)
        .join(Fattura, Fattura.location_id == Location.id)
        .where(Fattura.fornitore_id == fornitore_id)
    )
    if data_da:
        stmt = stmt.where(Fattura.data_documento >= data_da)
    if data_a:
        stmt = stmt.where(Fattura.data_documento <= data_a)
    stmt = stmt.distinct()
    locations = (await db.scalars(stmt)).all()
    
    # Questo è un mockup di astrazione in quanto richiede incrociare 
    # unità di misura in volumi per la specifica
    vendor_data = {
        "vendor_name": fornitore.nome_azienda,
        "location_servite": len(locations),
        "frequenza": "Settimanale (stimata)",
        "assorbimento": [
            {"categoria": "Volume Acquistato YTD (Mock)", "volume": 12000, "unita": "Pz"}
        ]
    }
    
    pdf_bytes = generate_vendor_passport_pdf(vendor_data)
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=Vendor_Passport_{fornitore_id}.pdf"
        }
    )


# ─────────────────────────────────────────────
# 1. Gestione Approvazioni Prezzi Manuali
# ─────────────────────────────────────────────

@router.post("/approvazioni", response_model=ApprovazionePrezzoResponse, summary="Crea o aggiorna approvazione prezzo")
async def create_approvazione(
    data: ApprovazionePrezzoCreate,
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    # Cerca se esiste gia'
    stmt = select(ApprovazionePrezzo).where(
        and_(
            ApprovazionePrezzo.sku_interno == data.sku_interno,
            ApprovazionePrezzo.descrizione_orig == data.descrizione_orig,
            ApprovazionePrezzo.mese == data.mese,
        )
    )
    res = await db.execute(stmt)
    appr = res.scalar_one_or_none()
    
    if appr:
        appr.prezzo_approvato = data.prezzo_approvato
        appr.stato = data.stato
    else:
        appr = ApprovazionePrezzo(
            sku_interno=data.sku_interno,
            descrizione_orig=data.descrizione_orig,
            mese=data.mese,
            prezzo_approvato=data.prezzo_approvato,
            stato=data.stato,
            created_at=datetime.now(timezone.utc),
        )
        db.add(appr)

    # ── SPEC §5.1: Aggiornamento Automatico ListinoMaster ──
    if data.stato.upper() in ("APPROVATO", "ACCETTATO"):
        # Cerca listino master attivo
        listino_stmt = select(ListinoMaster).where(
            and_(
                ListinoMaster.sku_interno == data.sku_interno,
                ListinoMaster.data_scadenza.is_(None)
            )
        ).limit(1)
        listino_res = await db.execute(listino_stmt)
        listino_active = listino_res.scalar_one_or_none()

        if listino_active:
            if float(listino_active.prezzo_pattuito) != float(data.prezzo_approvato):
                # Chiude la validita' del record corrente (Append-Only)
                listino_active.data_scadenza = date.today()
                
                # Crea nuovo record con prezzo aggiornato
                new_list = ListinoMaster(
                    fornitore_id=listino_active.fornitore_id,
                    sku_interno=data.sku_interno,
                    descrizione=listino_active.descrizione,
                    prezzo_pattuito=data.prezzo_approvato,
                    unita_misura=listino_active.unita_misura,
                    data_inizio_validita=date.today(),
                    data_scadenza=None
                )
                db.add(new_list)
        else:
            # Fallback: cerca una riga fattura recente per ottenere il fornitore_id e dettagli omologhi
            rf_stmt = (
                select(RigaFattura)
                .join(Fattura)
                .where(RigaFattura.sku_interno == data.sku_interno)
                .limit(1)
            )
            rf_res = await db.execute(rf_stmt)
            rf_item = rf_res.scalar_one_or_none()
            if rf_item and rf_item.fattura:
                new_list = ListinoMaster(
                    fornitore_id=rf_item.fattura.fornitore_id,
                    sku_interno=data.sku_interno,
                    descrizione=rf_item.descrizione_fornitore_raw or data.descrizione_orig,
                    prezzo_pattuito=data.prezzo_approvato,
                    unita_misura=rf_item.unita_misura_fattura or "Pz",
                    data_inizio_validita=date.today(),
                    data_scadenza=None
                )
                db.add(new_list)
        
    await db.flush()
    await db.refresh(appr)
    return appr


@router.get("/approvazioni", response_model=list[ApprovazionePrezzoResponse], summary="Lista approvazioni prezzi")
async def list_approvazioni(
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ApprovazionePrezzo).order_by(ApprovazionePrezzo.created_at.desc())
    res = await db.execute(stmt)
    return res.scalars().all()


@router.get("/price-trend/{sku_interno}", summary="Trend Storico Prezzi per SKU")
async def get_price_trend(
    sku_interno: str,
    _user = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Recupera lo storico degli acquisti cronologicamente
    stmt = (
        select(
            Fattura.data_documento,
            RigaFattura.prezzo_netto_normalizzato,
            RigaFattura.quantita,
            Fornitore.nome_azienda.label("fornitore_nome")
        )
        .join(RigaFattura, RigaFattura.fattura_id == Fattura.id)
        .join(Fornitore, Fattura.fornitore_id == Fornitore.id)
        .where(
            and_(
                RigaFattura.sku_interno == sku_interno,
                RigaFattura.stato_matching == "matched"
            )
        )
        .order_by(Fattura.data_documento.asc())
    )
    res = await db.execute(stmt)
    history = res.all()
    
    # Recupera il listino attivo corrente
    listino_stmt = select(ListinoMaster).where(
        and_(
            ListinoMaster.sku_interno == sku_interno,
            ListinoMaster.data_scadenza.is_(None)
        )
    ).limit(1)
    listino_res = await db.execute(listino_stmt)
    listino_active = listino_res.scalar_one_or_none()
    prezzo_contratto = float(listino_active.prezzo_pattuito) if listino_active else None

    points = []
    for r in history:
        points.append({
            "data": r.data_documento.isoformat() if isinstance(r.data_documento, date) else str(r.data_documento),
            "prezzo_pagato": float(r.prezzo_netto_normalizzato),
            "quantita": float(r.quantita),
            "fornitore": r.fornitore_nome,
            "prezzo_contratto": prezzo_contratto
        })
        
    return {
        "sku_interno": sku_interno,
        "prodotto_nome": listino_active.descrizione if listino_active else sku_interno,
        "prezzo_contratto_corrente": prezzo_contratto,
        "history": points
    }


# ─────────────────────────────────────────────
# 2. Audit & Anomalie Pricing (ConfrontoPrezzi)
# ─────────────────────────────────────────────

@router.get("/pricing-audit", summary="Audit e Rilevamento Anomalie di Pricing")
async def get_pricing_audit(
    location_id: int | None = Query(None),
    anno: str | None = Query(None),
    soglia: float = Query(5.0, description="Soglia percentuale minima di rincaro"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    sql = """
    WITH vendite_mensili AS (
        SELECT 
            rf.sku_interno as prodotto_id,
            rf.descrizione_fornitore_raw as nome_normalizzato,
            fo.nome_azienda as fornitore_ragione_sociale,
            f.fornitore_id,
            to_char(f.data_documento, 'YYYY-MM') as mese,
            AVG(rf.prezzo_netto_normalizzato) as prezzo_medio,
            SUM(rf.quantita) as qta_mese_corrente
        FROM righe_fattura rf
        JOIN fatture f ON rf.fattura_id = f.id
        JOIN fornitori fo ON f.fornitore_id = fo.id
        WHERE rf.sku_interno IS NOT NULL
          AND (:location_id IS NULL OR f.location_id = :location_id)
          AND (:anno IS NULL OR to_char(f.data_documento, 'YYYY') = :anno)
        GROUP BY rf.sku_interno, rf.descrizione_fornitore_raw, f.fornitore_id, fo.nome_azienda, to_char(f.data_documento, 'YYYY-MM')
    ),
    lagged_prezzi AS (
        SELECT 
            v.*,
            LAG(prezzo_medio, 1) OVER (PARTITION BY prodotto_id, fornitore_id ORDER BY mese) as prezzo_precedente_lag
        FROM vendite_mensili v
    )
    SELECT 
        lp.prodotto_id as sku_interno,
        lp.nome_normalizzato,
        lp.fornitore_id,
        lp.fornitore_ragione_sociale,
        lp.mese,
        lp.prezzo_medio,
        lp.qta_mese_corrente,
        lp.prezzo_precedente_lag,
        lc.prezzo_pattuito as prezzo_concordato,
        COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag) as prezzo_precedente,
        (SELECT MIN(rf2.prezzo_netto_normalizzato) FROM righe_fattura rf2 JOIN fatture f2 ON rf2.fattura_id = f2.id WHERE rf2.sku_interno = lp.prodotto_id AND rf2.prezzo_netto_normalizzato > 0 AND rf2.is_omaggio IS NOT TRUE) as hist_prezzo_min,
        (SELECT MAX(rf2.prezzo_netto_normalizzato) FROM righe_fattura rf2 JOIN fatture f2 ON rf2.fattura_id = f2.id WHERE rf2.sku_interno = lp.prodotto_id AND rf2.prezzo_netto_normalizzato > 0 AND rf2.is_omaggio IS NOT TRUE) as hist_prezzo_max,
        ap.stato
    FROM lagged_prezzi lp
    LEFT JOIN listino_master lc ON lp.prodotto_id = lc.sku_interno AND lp.fornitore_id = lc.fornitore_id
    LEFT JOIN approvazioni_prezzo ap ON lp.prodotto_id = ap.sku_interno AND lp.nome_normalizzato = ap.descrizione_orig AND lp.mese = ap.mese
    WHERE COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag) IS NOT NULL
      AND lp.prezzo_medio > COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)
      AND (((lp.prezzo_medio - COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) / COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) * 100) >= :soglia
    ORDER BY lp.mese DESC, lp.prodotto_id ASC
    LIMIT :limit OFFSET :offset
    """
    
    # Run total count query
    count_sql = f"""
    SELECT COUNT(*) FROM (
        WITH vendite_mensili AS (
            SELECT 
                rf.sku_interno as prodotto_id,
                rf.descrizione_fornitore_raw as nome_normalizzato,
                f.fornitore_id,
                to_char(f.data_documento, 'YYYY-MM') as mese,
                AVG(rf.prezzo_netto_normalizzato) as prezzo_medio
            FROM righe_fattura rf
            JOIN fatture f ON rf.fattura_id = f.id
            WHERE rf.sku_interno IS NOT NULL
              AND (:location_id IS NULL OR f.location_id = :location_id)
              AND (:anno IS NULL OR to_char(f.data_documento, 'YYYY') = :anno)
            GROUP BY rf.sku_interno, rf.descrizione_fornitore_raw, f.fornitore_id, to_char(f.data_documento, 'YYYY-MM')
        ),
        lagged_prezzi AS (
            SELECT 
                v.*,
                LAG(prezzo_medio, 1) OVER (PARTITION BY prodotto_id, fornitore_id ORDER BY mese) as prezzo_precedente_lag
            FROM vendite_mensili v
        )
        SELECT lp.prodotto_id
        FROM lagged_prezzi lp
        LEFT JOIN listino_master lc ON lp.prodotto_id = lc.sku_interno AND lp.fornitore_id = lc.fornitore_id
        WHERE COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag) IS NOT NULL
          AND lp.prezzo_medio > COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)
          AND (((lp.prezzo_medio - COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) / COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) * 100) >= :soglia
    ) as t
    """
    
    params = {
        "location_id": location_id,
        "anno": anno,
        "soglia": soglia,
    }
    
    # Counts
    count_res = await db.execute(text(count_sql), params)
    total = count_res.scalar() or 0
    
    params_with_limits = {**params, "limit": limit, "offset": offset}
    res = await db.execute(text(sql), params_with_limits)
    
    results = []
    for r in res.all():
        prezzo_medio = float(r.prezzo_medio)
        prezzo_precedente = float(r.prezzo_precedente)
        rincaro_unitario = prezzo_medio - prezzo_precedente
        spreco_mensile = rincaro_unitario * float(r.qta_mese_corrente)
        proiezione_annua = spreco_mensile * 12
        
        results.append({
            "sku_interno": r.sku_interno,
            "nome_normalizzato": r.nome_normalizzato,
            "fornitore_id": r.fornitore_id,
            "fornitore_ragione_sociale": r.fornitore_ragione_sociale,
            "mese": r.mese,
            "prezzo_medio": prezzo_medio,
            "qta_mese_corrente": float(r.qta_mese_corrente),
            "prezzo_precedente": prezzo_precedente,
            "rincaro_unitario": rincaro_unitario,
            "spreco_mensile": spreco_mensile,
            "proiezione_annua": proiezione_annua,
            "hist_prezzo_min": float(r.hist_prezzo_min) if r.hist_prezzo_min is not None else prezzo_medio,
            "hist_prezzo_max": float(r.hist_prezzo_max) if r.hist_prezzo_max is not None else prezzo_medio,
            "stato": r.stato or "PENDING"
        })
        
    return {"total": total, "results": results}


# ─────────────────────────────────────────────
# 3. Classifica Efficienza Locali
# ─────────────────────────────────────────────

@router.get("/efficiency-leaderboard", summary="Classifica Efficienza Locali")
async def get_efficiency_leaderboard(
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    sql = """
    WITH hist_min AS (
        SELECT 
            sku_interno,
            MIN(prezzo_netto_normalizzato) as hist_prezzo_min
        FROM righe_fattura
        WHERE sku_interno IS NOT NULL 
          AND stato_matching = 'matched'
          AND prezzo_netto_normalizzato > 0
          AND is_omaggio IS NOT TRUE
        GROUP BY sku_interno
    ),
    purchases AS (
        SELECT 
            f.location_id,
            loc.nome_struttura,
            rf.id as riga_id,
            rf.prezzo_netto_normalizzato,
            hm.hist_prezzo_min,
            CASE WHEN rf.prezzo_netto_normalizzato <= (hm.hist_prezzo_min * 1.05) THEN 1 ELSE 0 END as is_optimal
        FROM righe_fattura rf
        JOIN hist_min hm ON rf.sku_interno = hm.sku_interno
        JOIN fatture f ON rf.fattura_id = f.id
        JOIN location loc ON f.location_id = loc.id
        WHERE rf.stato_matching = 'matched'
    )
    SELECT 
        location_id,
        nome_struttura,
        COUNT(riga_id) as totali,
        SUM(is_optimal) as ottimali,
        ROUND((SUM(is_optimal)::numeric / COUNT(riga_id)::numeric) * 100, 2) as score
    FROM purchases
    GROUP BY location_id, nome_struttura
    ORDER BY score DESC
    """
    res = await db.execute(text(sql))
    
    leaderboard = []
    for i, r in enumerate(res.all()):
        rank = i + 1
        medal = "🏆" if rank == 1 else ("🥈" if rank == 2 else ("🥉" if rank == 3 else f"#{rank}"))
        leaderboard.append({
            "rank": rank,
            "medal": medal,
            "location_id": r.location_id,
            "nome_struttura": r.nome_struttura,
            "totali": r.totali,
            "ottimali": int(r.ottimali or 0),
            "score": float(r.score or 0)
        })
    return leaderboard


@router.get("/variance-loss", summary="Analisi Sprechi per Mancata Ottimizzazione")
async def get_variance_loss(
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcola lo spreco finanziario (variance loss) causato dagli acquisti effettuati a un prezzo 
    superiore al prezzo minimo storico registrato per ciascun articolo standard (SKU).
    """
    sql = """
    WITH min_prices AS (
        SELECT 
            sku_interno, 
            MIN(prezzo_netto_normalizzato) AS min_price
        FROM righe_fattura
        WHERE stato_matching = 'matched' AND sku_interno IS NOT NULL AND prezzo_netto_normalizzato > 0
        GROUP BY sku_interno
    )
    SELECT 
        r.sku_interno,
        COALESCE(MAX(lm.descrizione), MAX(r.descrizione_fornitore_raw), r.sku_interno) AS prodotto_nome,
        COALESCE(MAX(f.nome_azienda), 'ND') AS fornitore_nome,
        COUNT(r.id) AS numero_acquisti,
        SUM(r.quantita) AS quantita_totale,
        mp.min_price AS prezzo_minimo,
        AVG(r.prezzo_netto_normalizzato) AS prezzo_medio,
        SUM(GREATEST(0, r.prezzo_netto_normalizzato - mp.min_price) * r.quantita) AS spreco_totale
    FROM righe_fattura r
    JOIN min_prices mp ON r.sku_interno = mp.sku_interno
    LEFT JOIN listino_master lm ON lm.sku_interno = r.sku_interno AND lm.data_scadenza IS NULL
    LEFT JOIN fornitori f ON f.id = lm.fornitore_id
    WHERE r.stato_matching = 'matched'
    GROUP BY r.sku_interno, mp.min_price
    HAVING SUM(GREATEST(0, r.prezzo_netto_normalizzato - mp.min_price) * r.quantita) > 0
    ORDER BY spreco_totale DESC
    LIMIT 10;
    """
    res = await db.execute(text(sql))
    
    results = []
    for r in res.all():
        results.append({
            "sku_interno": r.sku_interno,
            "prodotto_nome": r.prodotto_nome,
            "fornitore_nome": r.fornitore_nome,
            "numero_acquisti": int(r.numero_acquisti or 0),
            "quantita_totale": float(r.quantita_totale or 0),
            "prezzo_minimo": float(r.prezzo_minimo or 0),
            "prezzo_medio": float(r.prezzo_medio or 0),
            "spreco_totale": float(r.spreco_totale or 0)
        })
    return results


# ─────────────────────────────────────────────
# 4. Esporta Excel di Contestazione (openpyxl)
# ─────────────────────────────────────────────

@router.get("/export-dispute-excel", summary="Esporta Excel Contestazione")
async def export_dispute_excel(
    location_id: int | None = Query(None),
    anno: str | None = Query(None),
    soglia: float = Query(5.0),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    sql = """
    WITH vendite_mensili AS (
        SELECT 
            rf.sku_interno as prodotto_id,
            rf.descrizione_fornitore_raw as nome_normalizzato,
            fo.nome_azienda as fornitore_ragione_sociale,
            f.fornitore_id,
            to_char(f.data_documento, 'YYYY-MM') as mese,
            AVG(rf.prezzo_netto_normalizzato) as prezzo_medio,
            SUM(rf.quantita) as qta_mese_corrente
        FROM righe_fattura rf
        JOIN fatture f ON rf.fattura_id = f.id
        JOIN fornitori fo ON f.fornitore_id = fo.id
        WHERE rf.sku_interno IS NOT NULL
          AND (:location_id IS NULL OR f.location_id = :location_id)
          AND (:anno IS NULL OR to_char(f.data_documento, 'YYYY') = :anno)
        GROUP BY rf.sku_interno, rf.descrizione_fornitore_raw, f.fornitore_id, fo.nome_azienda, to_char(f.data_documento, 'YYYY-MM')
    ),
    lagged_prezzi AS (
        SELECT 
            v.*,
            LAG(prezzo_medio, 1) OVER (PARTITION BY prodotto_id, fornitore_id ORDER BY mese) as prezzo_precedente_lag
        FROM vendite_mensili v
    )
    SELECT 
        lp.prodotto_id as sku_interno,
        lp.nome_normalizzato,
        lp.fornitore_id,
        lp.fornitore_ragione_sociale,
        lp.mese,
        lp.prezzo_medio,
        lp.qta_mese_corrente,
        lp.prezzo_precedente_lag,
        lc.prezzo_pattuito as prezzo_concordato,
        COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag) as prezzo_precedente
    FROM lagged_prezzi lp
    LEFT JOIN listino_master lc ON lp.prodotto_id = lc.sku_interno AND lp.fornitore_id = lc.fornitore_id
    WHERE COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag) IS NOT NULL
      AND lp.prezzo_medio > COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)
      AND (((lp.prezzo_medio - COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) / COALESCE(lc.prezzo_pattuito, lp.prezzo_precedente_lag)) * 100) >= :soglia
    ORDER BY lp.mese DESC
    """
    params = {
        "location_id": location_id,
        "anno": anno,
        "soglia": soglia,
    }
    
    res = await db.execute(text(sql), params)
    
    # Crea Workbook excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Rincari"
    
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Mese", "SKU Interno", "Prodotto", "Fornitore", 
        "Prezzo Target (€)", "Prezzo Medio Rilevato (€)", 
        "Quantità Acquistata", "Rincaro Unitario (€)", "Spreco Rilevato (€)"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # Aggiungi i dati
    row_num = 2
    for r in res.all():
        prezzo_medio = float(r.prezzo_medio)
        prezzo_precedente = float(r.prezzo_precedente)
        rincaro_unitario = prezzo_medio - prezzo_precedente
        quantita = float(r.qta_mese_corrente)
        spreco = rincaro_unitario * quantita
        
        row_data = [
            r.mese,
            r.sku_interno,
            r.nome_normalizzato,
            r.fornitore_ragione_sociale,
            prezzo_precedente,
            prezzo_medio,
            quantita,
            rincaro_unitario,
            spreco
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
            
            if col_idx in (5, 6, 8, 9):
                cell.number_format = '€ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 7:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
        row_num += 1
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="dossier_contestazione_rincari.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# ─────────────────────────────────────────────
# Report Consumi per Prodotto (Product Consumption)
# ─────────────────────────────────────────────

@router.get("/product-consumption", summary="Report Consumo per Prodotto")
async def get_product_consumption(
    location_ids: str | None = Query(None),
    fornitore_id: int | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Ritorna il report aggregato di consumo per ciascun SKU.
    """
    loc_ids = []
    if location_ids:
        try:
            loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        except ValueError:
            pass

    location_filter = ""
    params = {
        "fornitore_id": fornitore_id,
        "data_da": data_da,
        "data_a": data_a
    }
    if loc_ids:
        id_placeholders = ",".join(f":loc_id_{i}" for i in range(len(loc_ids)))
        location_filter = f"AND f.location_id IN ({id_placeholders})"
        for i, val in enumerate(loc_ids):
            params[f"loc_id_{i}"] = val

    sql = f"""
    SELECT 
        rf.sku_interno, 
        MAX(rf.descrizione_fornitore_raw) as descrizione,
        SUM(rf.quantita) as quantita_totale,
        SUM(CASE WHEN rf.is_omaggio = TRUE THEN rf.quantita ELSE 0 END) as quantita_omaggio,
        COALESCE(
            MAX(CASE WHEN rf.is_omaggio = FALSE AND rf.unita_misura_fattura NOT IN ('OMAGGIO', 'omaggio', 'Omaggio') THEN rf.unita_misura_fattura END),
            MAX(rf.unita_misura_fattura)
        ) as unita_misura,
        SUM(rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale,
        CASE 
            WHEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END) > 0 
            THEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato * rf.quantita ELSE 0 END) / SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END)
            ELSE 0 
        END as prezzo_medio
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    WHERE rf.sku_interno IS NOT NULL
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    GROUP BY rf.sku_interno
    ORDER BY spesa_totale DESC
    """
    
    res = await db.execute(text(sql), params)
    
    results = []
    for r in res.all():
        results.append({
            "sku_interno": r.sku_interno,
            "descrizione": r.descrizione,
            "quantita_totale": float(r.quantita_totale or 0),
            "quantita_omaggio": float(r.quantita_omaggio or 0),
            "unita_misura": r.unita_misura or "Pz",
            "spesa_totale": float(r.spesa_totale or 0),
            "prezzo_medio": float(r.prezzo_medio or 0)
        })
    return results


@router.get("/product-consumption/{sku_interno}", summary="Dettaglio Consumo SKU per Location e Mese")
async def get_product_consumption_detail(
    sku_interno: str,
    location_ids: str | None = Query(None),
    fornitore_id: int | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    # Normalizza i parametri in caso di chiamata diretta in Python (es. unit tests)
    if not isinstance(location_ids, str):
        location_ids = None
    if not isinstance(fornitore_id, int):
        fornitore_id = None
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None

    skus = [x.strip() for x in sku_interno.split(",") if x.strip()]
    if not skus:
        return {
            "sku_interno": sku_interno,
            "consumo_per_location": [],
            "consumo_per_mese": []
        }

    loc_ids = []
    if location_ids:
        try:
            loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        except ValueError:
            pass

    location_filter = ""
    params = {
        "fornitore_id": fornitore_id,
        "data_da": data_da,
        "data_a": data_a
    }
    
    # Costruiamo il filtro SKU dinamico con bind variables
    sku_placeholders = ",".join(f":sku_{i}" for i in range(len(skus)))
    sku_filter = f"rf.sku_interno IN ({sku_placeholders})"
    for i, s in enumerate(skus):
        params[f"sku_{i}"] = s

    if loc_ids:
        id_placeholders = ",".join(f":loc_id_{i}" for i in range(len(loc_ids)))
        location_filter = f"AND f.location_id IN ({id_placeholders})"
        for i, val in enumerate(loc_ids):
            params[f"loc_id_{i}"] = val

    # 1. Split by Location
    sql_loc = f"""
    SELECT 
        l.nome_struttura as location_nome,
        SUM(rf.quantita) as quantita_totale,
        SUM(rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    JOIN location l ON f.location_id = l.id
    WHERE {sku_filter}
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    GROUP BY l.nome_struttura
    ORDER BY spesa_totale DESC
    """
    res_loc = await db.execute(text(sql_loc), params)
    by_location = []
    for r in res_loc.all():
        by_location.append({
            "location_nome": r.location_nome,
            "quantita_totale": float(r.quantita_totale or 0),
            "spesa_totale": float(r.spesa_totale or 0)
        })

    # 2. Split by Month
    sql_month = f"""
    SELECT 
        to_char(f.data_documento, 'YYYY-MM') as mese,
        SUM(rf.quantita) as quantita_totale,
        SUM(rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    WHERE {sku_filter}
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    GROUP BY to_char(f.data_documento, 'YYYY-MM')
    ORDER BY mese DESC
    """
    res_month = await db.execute(text(sql_month), params)
    by_month = []
    for r in res_month.all():
        by_month.append({
            "mese": r.mese,
            "quantita_totale": float(r.quantita_totale or 0),
            "spesa_totale": float(r.spesa_totale or 0)
        })

    # 3. Aggregated Prices (Min, Max, Avg)
    sql_prices = f"""
    SELECT 
        MIN(CASE WHEN rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato END) as prezzo_minimo,
        MAX(CASE WHEN rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato END) as prezzo_massimo,
        CASE 
            WHEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END) > 0 
            THEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato * rf.quantita ELSE 0 END) / SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END)
            ELSE 0 
        END as prezzo_medio
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    WHERE {sku_filter}
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    """
    res_prices = await db.execute(text(sql_prices), params)
    row_prices = res_prices.one_or_none()
    prezzo_minimo = float(row_prices.prezzo_minimo or 0) if row_prices and row_prices.prezzo_minimo else 0.0
    prezzo_massimo = float(row_prices.prezzo_massimo or 0) if row_prices and row_prices.prezzo_massimo else 0.0
    prezzo_medio = float(row_prices.prezzo_medio or 0) if row_prices and row_prices.prezzo_medio else 0.0

    return {
        "sku_interno": sku_interno,
        "consumo_per_location": by_location,
        "consumo_per_mese": by_month,
        "prezzo_minimo": prezzo_minimo,
        "prezzo_massimo": prezzo_massimo,
        "prezzo_medio": prezzo_medio
    }


@router.get("/product-consumption/{sku_interno}/invoices", summary="Dettaglio Fatture che generano il Consumo dello SKU")
async def get_product_consumption_invoices(
    sku_interno: str,
    location_ids: str | None = Query(None),
    fornitore_id: int | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    if not isinstance(location_ids, str):
        location_ids = None
    if not isinstance(fornitore_id, int):
        fornitore_id = None
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None

    skus = [x.strip() for x in sku_interno.split(",") if x.strip()]
    if not skus:
        return []

    loc_ids = []
    if location_ids:
        try:
            loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        except ValueError:
            pass

    location_filter = ""
    params = {
        "fornitore_id": fornitore_id,
        "data_da": data_da,
        "data_a": data_a
    }
    
    sku_placeholders = ",".join(f":sku_{i}" for i in range(len(skus)))
    sku_filter = f"rf.sku_interno IN ({sku_placeholders})"
    for i, s in enumerate(skus):
        params[f"sku_{i}"] = s

    if loc_ids:
        id_placeholders = ",".join(f":loc_id_{i}" for i in range(len(loc_ids)))
        location_filter = f"AND f.location_id IN ({id_placeholders})"
        for i, val in enumerate(loc_ids):
            params[f"loc_id_{i}"] = val

    sql = f"""
    SELECT 
        f.id as fattura_id,
        f.numero_documento,
        f.data_documento,
        l.nome_struttura as location_nome,
        fo.nome_azienda as fornitore_nome,
        rf.descrizione_fornitore_raw as prodotto_descrizione,
        rf.quantita as quantita,
        rf.unita_misura_fattura as unita_misura,
        rf.prezzo_netto_normalizzato as prezzo_unitario,
        (rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale,
        rf.is_omaggio as is_omaggio,
        rf.sku_interno as sku_interno
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    JOIN location l ON f.location_id = l.id
    JOIN fornitori fo ON f.fornitore_id = fo.id
    WHERE {sku_filter}
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    ORDER BY f.data_documento DESC, f.numero_documento DESC
    """
    
    res = await db.execute(text(sql), params)
    results = []
    for r in res.all():
        results.append({
            "fattura_id": r.fattura_id,
            "numero_documento": r.numero_documento,
            "data_documento": r.data_documento.isoformat() if r.data_documento else None,
            "location_nome": r.location_nome,
            "fornitore_nome": r.fornitore_nome,
            "prodotto_descrizione": r.prodotto_descrizione,
            "quantita": float(r.quantita or 0),
            "unita_misura": r.unita_misura or "Pz",
            "prezzo_unitario": float(r.prezzo_unitario or 0),
            "spesa_totale": float(r.spesa_totale or 0),
            "is_omaggio": bool(r.is_omaggio or False),
            "sku_interno": r.sku_interno
        })
    return results


@router.get("/product-consumption/{sku_interno}/invoices-pdf", summary="Genera PDF del Riepilogo Fatture di Consumo dello/degli SKU")
async def get_product_consumption_invoices_pdf(
    sku_interno: str,
    location_ids: str | None = Query(None),
    fornitore_id: int | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    if not isinstance(location_ids, str):
        location_ids = None
    if not isinstance(fornitore_id, int):
        fornitore_id = None
    if not isinstance(data_da, date):
        data_da = None
    if not isinstance(data_a, date):
        data_a = None

    # Recupera i dati delle fatture
    invoices = await get_product_consumption_invoices(
        sku_interno=sku_interno,
        location_ids=location_ids,
        fornitore_id=fornitore_id,
        data_da=data_da,
        data_a=data_a,
        _admin=_admin,
        db=db
    )

    # Determina il titolo del report
    skus = [x.strip() for x in sku_interno.split(",") if x.strip()]
    
    title = sku_interno
    if len(skus) == 1:
        sql_desc = "SELECT descrizione_fornitore_raw FROM righe_fattura WHERE sku_interno = :sku LIMIT 1"
        res_desc = await db.execute(text(sql_desc), {"sku": skus[0]})
        row_desc = res_desc.first()
        if row_desc and row_desc[0]:
            title = f"{skus[0]} - {row_desc[0]}"
    else:
        title = f"Consolidato di {len(skus)} articoli ({', '.join(skus[:3])}{'...' if len(skus) > 3 else ''})"

    # Costruiamo una descrizione dei filtri
    filters_parts = []
    if location_ids:
        loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        if loc_ids:
            res_locs = await db.execute(text("SELECT nome_struttura FROM location WHERE id = ANY(:ids)"), {"ids": list(loc_ids)})
            names = [r[0] for r in res_locs.all()]
            filters_parts.append(f"Sedi: {', '.join(names)}")
    if fornitore_id:
        res_forn = await db.execute(text("SELECT nome_azienda FROM fornitori WHERE id = :fid"), {"fid": fornitore_id})
        row_forn = res_forn.first()
        if row_forn:
            filters_parts.append(f"Fornitore: {row_forn[0]}")
    if data_da:
        filters_parts.append(f"Da: {data_da.strftime('%d/%m/%Y')}")
    if data_a:
        filters_parts.append(f"A: {data_a.strftime('%d/%m/%Y')}")

    filter_desc = " | ".join(filters_parts) if filters_parts else "Nessuno"

    pdf_bytes = generate_consumption_invoices_pdf(title, invoices, filter_desc)
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=Riepilogo_Fatture_Consumo.pdf"
        }
    )


@router.get("/export-product-consumption-excel", summary="Esporta Excel Consumo per Prodotto")
async def export_product_consumption_excel(
    location_ids: str | None = Query(None),
    fornitore_id: int | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Genera ed esporta un report Excel (.xlsx) dei consumi per prodotto.
    """
    loc_ids = []
    if location_ids:
        try:
            loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        except ValueError:
            pass

    location_filter = ""
    params = {
        "fornitore_id": fornitore_id,
        "data_da": data_da,
        "data_a": data_a
    }
    if loc_ids:
        id_placeholders = ",".join(f":loc_id_{i}" for i in range(len(loc_ids)))
        location_filter = f"AND f.location_id IN ({id_placeholders})"
        for i, val in enumerate(loc_ids):
            params[f"loc_id_{i}"] = val

    sql = f"""
    SELECT 
        rf.sku_interno, 
        MAX(rf.descrizione_fornitore_raw) as descrizione,
        SUM(rf.quantita) as quantita_totale,
        SUM(CASE WHEN rf.is_omaggio = TRUE THEN rf.quantita ELSE 0 END) as quantita_omaggio,
        COALESCE(
            MAX(CASE WHEN rf.is_omaggio = FALSE AND rf.unita_misura_fattura NOT IN ('OMAGGIO', 'omaggio', 'Omaggio') THEN rf.unita_misura_fattura END),
            MAX(rf.unita_misura_fattura)
        ) as unita_misura,
        SUM(rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale,
        CASE 
            WHEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END) > 0 
            THEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato * rf.quantita ELSE 0 END) / SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END)
            ELSE 0 
        END as prezzo_medio
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    WHERE rf.sku_interno IS NOT NULL
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    GROUP BY rf.sku_interno
    ORDER BY spesa_totale DESC
    """
    
    res = await db.execute(text(sql), params)
    
    # Crea Workbook excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumi per Prodotto"
    
    # Mostra griglia
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "SKU Interno", "Prodotto", "Quantità Totale", "di cui Omaggi",
        "Unità di Misura", "Prezzo Medio (€)", "Spesa Totale (€)"
    ]
    
    ws.append(headers)
    
    # Stile intestazione elegante (Tonalità Navy coordinata a Price Sentinel)
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # Aggiungi i dati
    row_num = 2
    for r in res.all():
        quantita = float(r.quantita_totale or 0)
        quantita_omaggio = float(r.quantita_omaggio or 0)
        prezzo_medio = float(r.prezzo_medio or 0)
        spesa_totale = float(r.spesa_totale or 0)
        
        row_data = [
            r.sku_interno,
            r.descrizione,
            quantita,
            quantita_omaggio,
            r.unita_misura or "Pz",
            prezzo_medio,
            spesa_totale
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
            
            if col_idx in (6, 7):
                cell.number_format = '€ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3, 4):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (1, 5):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
        row_num += 1
        
    # Auto-fit colonne
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.column in (6, 7) and type(cell.value) in (int, float):
                val_str = f"€ {cell.value:.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="report_consumo_prodotti.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.get("/top-purchased-products", summary="Ottiene i prodotti più acquistati")
async def get_top_purchased_products(
    limit: int | None = Query(50, ge=1, le=1000),
    sort_by: str = Query("quantita", description="Ordinamento: quantita, spesa, acquisti"),
    fornitore_id: int | None = Query(None),
    location_ids: str | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Ritorna la lista dei prodotti più acquistati filtrati per fornitore, location e date range.
    """
    loc_ids = []
    if location_ids:
        try:
            loc_ids = [int(x) for x in location_ids.split(",") if x.strip()]
        except ValueError:
            pass

    location_filter = ""
    params = {
        "fornitore_id": fornitore_id,
        "data_da": data_da,
        "data_a": data_a,
        "limit": limit
    }
    if loc_ids:
        id_placeholders = ",".join(f":loc_id_{i}" for i in range(len(loc_ids)))
        location_filter = f"AND f.location_id IN ({id_placeholders})"
        for i, val in enumerate(loc_ids):
            params[f"loc_id_{i}"] = val

    valid_sorts = {
        "quantita": "quantita_totale DESC",
        "spesa": "spesa_totale DESC",
        "acquisti": "numero_acquisti DESC"
    }
    order_by_clause = valid_sorts.get(sort_by, "quantita_totale DESC")

    sql = f"""
    SELECT 
        rf.sku_interno, 
        MAX(rf.descrizione_fornitore_raw) as descrizione,
        SUM(rf.quantita) as quantita_totale,
        SUM(CASE WHEN rf.is_omaggio = TRUE THEN rf.quantita ELSE 0 END) as quantita_omaggio,
        COALESCE(
            MAX(CASE WHEN rf.is_omaggio = FALSE AND rf.unita_misura_fattura NOT IN ('OMAGGIO', 'omaggio', 'Omaggio') THEN rf.unita_misura_fattura END),
            MAX(rf.unita_misura_fattura)
        ) as unita_misura,
        SUM(rf.prezzo_netto_normalizzato * rf.quantita) as spesa_totale,
        COUNT(rf.id) as numero_acquisti,
        CASE 
            WHEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END) > 0 
            THEN SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.prezzo_netto_normalizzato * rf.quantita ELSE 0 END) / SUM(CASE WHEN rf.is_omaggio = FALSE AND rf.prezzo_netto_normalizzato > 0 THEN rf.quantita ELSE 0 END)
            ELSE 0 
        END as prezzo_medio
    FROM righe_fattura rf
    JOIN fatture f ON rf.fattura_id = f.id
    WHERE rf.sku_interno IS NOT NULL
      {location_filter}
      AND (cast(:fornitore_id as integer) IS NULL OR f.fornitore_id = cast(:fornitore_id as integer))
      AND (cast(:data_da as date) IS NULL OR f.data_documento >= cast(:data_da as date))
      AND (cast(:data_a as date) IS NULL OR f.data_documento <= cast(:data_a as date))
    GROUP BY rf.sku_interno
    ORDER BY {order_by_clause}
    LIMIT :limit
    """
    
    res = await db.execute(text(sql), params)
    
    results = []
    for r in res.all():
        results.append({
            "sku_interno": r.sku_interno,
            "descrizione": r.descrizione,
            "quantita_totale": float(r.quantita_totale or 0),
            "quantita_omaggio": float(r.quantita_omaggio or 0),
            "unita_misura": r.unita_misura or "Pz",
            "spesa_totale": float(r.spesa_totale or 0),
            "numero_acquisti": int(r.numero_acquisti or 0),
            "prezzo_medio": float(r.prezzo_medio or 0)
        })
    return results


@router.get("/export-top-purchased-excel", summary="Esporta Excel dei prodotti più acquistati")
async def export_top_purchased_excel(
    limit: int | None = Query(50, ge=1, le=1000),
    sort_by: str = Query("quantita", description="Ordinamento: quantita, spesa, acquisti"),
    fornitore_id: int | None = Query(None),
    location_ids: str | None = Query(None),
    data_da: date | None = Query(None),
    data_a: date | None = Query(None),
    _admin = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Genera ed esporta un report Excel (.xlsx) dei prodotti più acquistati con filtri.
    """
    data = await get_top_purchased_products(
        limit=limit,
        sort_by=sort_by,
        fornitore_id=fornitore_id,
        location_ids=location_ids,
        data_da=data_da,
        data_a=data_a,
        _admin=_admin,
        db=db
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listino Top Prodotti"
    
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Posizione", "SKU Interno", "Prodotto", "Unità di Misura", 
        "Quantità Totale", "Numero Acquisti", "Prezzo Medio (€)", "Spesa Totale (€)"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    row_num = 2
    for idx, item in enumerate(data):
        row_data = [
            idx + 1,
            item["sku_interno"],
            item["descrizione"],
            item["unita_misura"],
            item["quantita_totale"],
            item["numero_acquisti"],
            item["prezzo_medio"],
            item["spesa_totale"]
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
            
            if col_idx in (7, 8):
                cell.number_format = '€ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (1, 2, 4):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
        row_num += 1
        
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.column in (7, 8) and type(cell.value) in (int, float):
                val_str = f"€ {cell.value:.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"listino_top_{limit or 'all'}_prodotti.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )



