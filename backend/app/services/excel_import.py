"""
Price Sentinel — Excel Import/Export Service.
Import Listino Master da file Excel (template standard).
Spec §8 - Sprint 1: template Excel listini.
"""

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# Template Excel per Listino Master
# ─────────────────────────────────────────────

TEMPLATE_COLUMNS = [
    ("sku_interno", "Codice SKU Interno", 20),
    ("descrizione", "Descrizione Prodotto", 40),
    ("prezzo_pattuito", "Prezzo Pattuito (€)", 18),
    ("unita_misura", "Unità di Misura", 16),
    ("data_inizio_validita", "Data Inizio Validità", 20),
    ("pfa_tipo", "PFA Tipo (Percentuale/Fisso/Scaglioni)", 30),
    ("pfa_valore", "PFA Valore", 14),
]

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EXAMPLE_FILL = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")


def generate_template_excel(fornitore_nome: str = "NomeFornitore") -> bytes:
    """
    Genera un file Excel template per l'import del Listino Master.
    Restituisce i bytes del file .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Listino Master"

    # ── Istruzioni ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"📋 PRICE SENTINEL — Template Listino: {fornitore_nome}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1B2A4A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Compila le righe a partire dalla riga 5. "
        "Non modificare l'intestazione (riga 4). "
        "Formato date: GG/MM/AAAA. "
        "PFA Tipo: Percentuale, Fisso, Scaglioni, oppure vuoto."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    # ── Riga vuota ──
    ws.row_dimensions[3].height = 5

    # ── Header ──
    for col_idx, (_, label, width) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 25

    # ── Riga esempio ──
    example_data = [
        "GIN-HENDRICKS-70CL",
        "Gin Hendrick's 70cl",
        "22.5000",
        "Pz",
        "01/01/2025",
        "Percentuale",
        "0.0200",
    ]
    for col_idx, val in enumerate(example_data, start=1):
        cell = ws.cell(row=5, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=11, italic=True, color="999999")
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER

    # ── Congela pannelli ──
    ws.freeze_panes = "A5"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A4:G4"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─────────────────────────────────────────────
# Parser / Validator Excel → Listino Records
# ─────────────────────────────────────────────

class ExcelValidationError:
    """Singolo errore di validazione."""
    def __init__(self, row: int, column: str, message: str):
        self.row = row
        self.column = column
        self.message = message

    def to_dict(self):
        return {"row": self.row, "column": self.column, "message": self.message}


class ExcelParseResult:
    """Risultato del parsing: record validi + errori."""
    def __init__(self):
        self.records: list[dict] = []
        self.errors: list[ExcelValidationError] = []
        self.total_rows: int = 0

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    def to_dict(self):
        return {
            "total_rows": self.total_rows,
            "valid_records": len(self.records),
            "errors_count": len(self.errors),
            "errors": [e.to_dict() for e in self.errors],
        }


PFA_VALID_TYPES = {"percentuale", "fisso", "scaglioni", ""}


def parse_listino_excel(file_data: BinaryIO, fornitore_id: int) -> ExcelParseResult:
    """
    Parsa un file Excel e restituisce una lista di record validi
    per l'inserimento nel ListinoMaster.

    Validazioni:
    - sku_interno: obbligatorio, max 100 char
    - descrizione: obbligatoria
    - prezzo_pattuito: obbligatorio, NUMERIC(12,4), > 0
    - unita_misura: obbligatoria (Kg, Lt, Pz, Cassa, ecc.)
    - data_inizio_validita: obbligatoria, formato date
    - pfa_tipo: opzionale (Percentuale|Fisso|Scaglioni)
    - pfa_valore: obbligatorio se pfa_tipo è Percentuale o Fisso
    """
    result = ExcelParseResult()

    try:
        wb = load_workbook(file_data, read_only=True, data_only=True)
    except Exception:
        result.errors.append(
            ExcelValidationError(0, "-", "File non valido. Usa il template .xlsx fornito.")
        )
        return result

    ws = wb.active

    # Trova la riga header (cerchiamo "Codice SKU" nella prima colonna)
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=1, values_only=False), start=1):
        cell_val = str(row[0].value or "").strip().lower()
        if "codice" in cell_val or "sku" in cell_val:
            header_row = row_idx
            break

    if header_row is None:
        # Fallback: assume riga 4 (template standard)
        header_row = 4

    data_start = header_row + 1

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, max_col=7, values_only=True),
        start=data_start,
    ):
        # Salta righe completamente vuote
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        result.total_rows += 1
        sku, desc, prezzo, uom, data_val, pfa_tipo, pfa_valore = (
            row[i] if i < len(row) else None for i in range(7)
        )

        record = {"fornitore_id": fornitore_id}
        row_errors = False

        # ── SKU ──
        sku_str = str(sku or "").strip()
        if not sku_str:
            result.errors.append(ExcelValidationError(row_idx, "A (SKU)", "SKU obbligatorio"))
            row_errors = True
        elif len(sku_str) > 100:
            result.errors.append(ExcelValidationError(row_idx, "A (SKU)", "SKU max 100 caratteri"))
            row_errors = True
        else:
            record["sku_interno"] = sku_str

        # ── Descrizione ──
        desc_str = str(desc or "").strip()
        if not desc_str:
            result.errors.append(ExcelValidationError(row_idx, "B (Descrizione)", "Descrizione obbligatoria"))
            row_errors = True
        else:
            record["descrizione"] = desc_str

        # ── Prezzo ──
        try:
            if prezzo is None or str(prezzo).strip() == "":
                raise ValueError("vuoto")
            prezzo_dec = Decimal(str(prezzo).strip().replace(",", "."))
            if prezzo_dec <= 0:
                raise ValueError("negativo")
            if prezzo_dec >= Decimal("100000000"):
                raise ValueError("troppo grande")
            record["prezzo_pattuito"] = prezzo_dec
        except (InvalidOperation, ValueError) as e:
            result.errors.append(
                ExcelValidationError(row_idx, "C (Prezzo)", f"Prezzo non valido: {e}")
            )
            row_errors = True

        # ── UoM ──
        uom_str = str(uom or "").strip()
        if not uom_str:
            result.errors.append(ExcelValidationError(row_idx, "D (UoM)", "Unità di misura obbligatoria"))
            row_errors = True
        else:
            record["unita_misura"] = uom_str

        # ── Data Inizio Validità ──
        if isinstance(data_val, datetime):
            record["data_inizio_validita"] = data_val.date()
        elif isinstance(data_val, date):
            record["data_inizio_validita"] = data_val
        elif isinstance(data_val, str):
            data_str = data_val.strip()
            parsed = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    parsed = datetime.strptime(data_str, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed:
                record["data_inizio_validita"] = parsed
            else:
                result.errors.append(
                    ExcelValidationError(row_idx, "E (Data)", f"Formato data non riconosciuto: '{data_str}'")
                )
                row_errors = True
        else:
            result.errors.append(
                ExcelValidationError(row_idx, "E (Data)", "Data inizio validità obbligatoria")
            )
            row_errors = True

        # ── PFA Tipo ──
        pfa_tipo_str = str(pfa_tipo or "").strip().lower()
        if pfa_tipo_str and pfa_tipo_str not in PFA_VALID_TYPES:
            result.errors.append(
                ExcelValidationError(
                    row_idx, "F (PFA Tipo)",
                    f"Tipo PFA non valido: '{pfa_tipo_str}'. Usa: Percentuale, Fisso, Scaglioni"
                )
            )
            row_errors = True
        else:
            record["pfa_tipo"] = pfa_tipo_str.capitalize() if pfa_tipo_str else None

        # ── PFA Valore ──
        if pfa_tipo_str in ("percentuale", "fisso"):
            try:
                if pfa_valore is None or str(pfa_valore).strip() == "":
                    raise ValueError("vuoto")
                pfa_val_dec = Decimal(str(pfa_valore).strip().replace(",", "."))
                record["pfa_valore"] = pfa_val_dec
            except (InvalidOperation, ValueError):
                result.errors.append(
                    ExcelValidationError(
                        row_idx, "G (PFA Valore)",
                        f"PFA Valore obbligatorio per tipo '{pfa_tipo_str}'"
                    )
                )
                row_errors = True
        else:
            record["pfa_valore"] = None

        if not row_errors:
            result.records.append(record)

    wb.close()
    return result
