import re
import io
import hashlib
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Optional, Any
from openpyxl import load_workbook

# Stable SKU generation from product description
def generate_stable_sku(descrizione: str) -> str:
    """
    Genera uno SKU interno stabile, univoco e pulito a partire dalla descrizione.
    Es: 'POLIPO MAROCCO T6 BLOCCO' -> 'POLIPO-MAROCCO-T6-BLOCCO'
    """
    s = descrizione.upper().strip()
    # Mantieni solo lettere, cifre, spazi, slash, punti e trattini
    s = re.sub(r'[^A-Z0-9\s/.-]', '', s)
    # Rimpiazza spazi, slash e punti con trattini
    s = re.sub(r'[\s/.]+', '-', s)
    # Evita trattini consecutivi
    s = re.sub(r'-+', '-', s)
    return s.strip('-')

# Deterministic P.IVA generation to avoid conflicts
def generate_deterministic_piva(supplier_name: str) -> str:
    """
    Genera una Partita IVA fittizia di 11 cifre basata sull'hash del nome del fornitore.
    Questo garantisce che lo stesso nome generi sempre la stessa P.IVA fittizia.
    """
    h = hashlib.sha256(supplier_name.encode('utf-8')).hexdigest()
    # Estrae solo i caratteri numerici
    digits = ''.join(c for c in h if c.isdigit())
    if len(digits) < 11:
        digits = (digits + "12345678901")[:11]
    else:
        digits = digits[:11]
    return digits

class MultiExcelValidationError:
    def __init__(self, row: int, column: str, message: str):
        self.row = row
        self.column = column
        self.message = message

    def to_dict(self):
        return {"row": self.row, "column": self.column, "message": self.message}

class MultiExcelParseResult:
    def __init__(self):
        self.records: list[dict] = []
        self.errors: list[MultiExcelValidationError] = []
        self.suppliers_detected: list[str] = []
        self.total_rows: int = 0

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    def to_dict(self):
        return {
            "total_rows": self.total_rows,
            "valid_records": len(self.records),
            "suppliers_detected": self.suppliers_detected,
            "errors_count": len(self.errors),
            "errors": [e.to_dict() for e in self.errors],
        }

def parse_multi_supplier_excel(file_data: BinaryIO) -> MultiExcelParseResult:
    """
    Parsa un listino comparativo multi-fornitore Excel.
    Colonne previste:
    PRODOTTO | UNITA DI MISURA | FORNITORE 1 | FORNITORE 2 | ... | MIGLIOR FORNITORE (opzionale)
    """
    result = MultiExcelParseResult()

    try:
        # Prevent XML Entity Expansion vulnerabilities in openpyxl by using default load settings
        # which are secure in standard openpyxl versions.
        wb = load_workbook(file_data, read_only=True, data_only=True)
    except Exception:
        result.errors.append(
            MultiExcelValidationError(0, "-", "File non valido. Assicurati che sia un file Excel .xlsx corretto.")
        )
        return result

    ws = wb.active

    # 1. Trova la riga intestazione
    header_row_idx = None
    headers: list[str] = []
    
    # Cerchiamo nelle prime 15 righe una riga contenente "prodotto" o "descrizione"
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        row_clean = [str(c).lower().strip() if c is not None else "" for c in row]
        if any("prodotto" in s or "descrizione" in s for s in row_clean):
            header_row_idx = row_idx
            headers = [str(c).strip() if c is not None else "" for c in row]
            break

    if header_row_idx is None:
        result.errors.append(
            MultiExcelValidationError(0, "-", "Intestazione non trovata. Il file deve contenere una colonna 'PRODOTTO' o 'DESCRIZIONE'.")
        )
        wb.close()
        return result

    # 2. Mappa le colonne
    prodotto_idx = None
    uom_idx = None
    fornitore_scelto_idx = None
    
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        if not h:
            continue
        if "prodotto" in h_lower or "descrizione" in h_lower:
            if prodotto_idx is None:
                prodotto_idx = idx
        elif any(k in h_lower for k in ("peso", "unita", "unità", "conf", "uom")):
            if uom_idx is None:
                uom_idx = idx
        elif any(k in h_lower for k in ("scelto", "miglior", "attivo", "preferito")):
            if fornitore_scelto_idx is None:
                fornitore_scelto_idx = idx

    if prodotto_idx is None:
        result.errors.append(
            MultiExcelValidationError(header_row_idx, "-", "Impossibile identificare la colonna dei Prodotti.")
        )
        wb.close()
        return result

    # Trova le colonne dei fornitori
    supplier_cols: list[tuple[int, str]] = [] # list of (col_index, supplier_name)
    for idx, h in enumerate(headers):
        if not h:
            continue
        if idx == prodotto_idx:
            continue
        if idx == uom_idx:
            continue
        if idx == fornitore_scelto_idx:
            continue
        supplier_cols.append((idx, h))
        result.suppliers_detected.append(h)

    if not supplier_cols:
        result.errors.append(
            MultiExcelValidationError(header_row_idx, "-", "Nessun fornitore identificato nelle colonne.")
        )
        wb.close()
        return result

    # 3. Leggi i dati a partire dalla riga successiva all'intestazione
    data_start = header_row_idx + 1

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True),
        start=data_start,
    ):
        # Salta righe vuote
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        result.total_rows += 1
        
        # Prodotto descrizione
        prod_val = row[prodotto_idx] if prodotto_idx < len(row) else None
        prod_desc = str(prod_val).strip() if prod_val is not None else ""
        
        if not prod_desc:
            result.errors.append(
                MultiExcelValidationError(row_idx, f"Colonna {prodotto_idx + 1}", "Descrizione prodotto vuota.")
            )
            continue

        # UoM (Unita di Misura)
        uom_val = row[uom_idx] if (uom_idx is not None and uom_idx < len(row)) else None
        uom_str = str(uom_val).strip() if uom_val is not None else "Pz"
        if not uom_str:
            uom_str = "Pz"

        sku = generate_stable_sku(prod_desc)

        # Estrai i prezzi dei fornitori
        prezzi_fornitori = {}
        row_has_valid_price = False
        
        for col_idx, supplier_name in supplier_cols:
            price_val = row[col_idx] if col_idx < len(row) else None
            if price_val is None or str(price_val).strip() == "":
                continue
                
            price_str = str(price_val).strip().replace(",", ".")
            try:
                price_dec = Decimal(price_str)
                if price_dec <= 0:
                    result.errors.append(
                        MultiExcelValidationError(row_idx, supplier_name, f"Il prezzo deve essere positivo (trovato: {price_str})")
                    )
                    continue
                if price_dec >= Decimal("100000000"):
                    result.errors.append(
                        MultiExcelValidationError(row_idx, supplier_name, f"Prezzo eccessivo (trovato: {price_str})")
                    )
                    continue
                prezzi_fornitori[supplier_name] = price_dec
                row_has_valid_price = True
            except (InvalidOperation, ValueError):
                result.errors.append(
                    MultiExcelValidationError(row_idx, supplier_name, f"Formato prezzo non valido: '{price_val}'")
                )

        if row_has_valid_price:
            result.records.append({
                "sku_interno": sku,
                "descrizione": prod_desc,
                "unita_misura": uom_str,
                "prezzi": prezzi_fornitori
            })

    wb.close()
    return result
