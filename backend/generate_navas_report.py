import asyncio
import sys
from decimal import Decimal
from datetime import date

sys.path.append("/app")

from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

from app.models.fatture import RigaFattura, Fattura, StatoMatching
from app.models.listino import ListinoMaster
from app.models.fornitori import Fornitore

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATABASE_URL = "postgresql+asyncpg://sentinel:sentinel_dev_2025@db:5432/price_sentinel"
OUTPUT_PATH = "/app/storico_prezzi_navas.xlsx"

async def main():
    engine = create_async_engine(DATABASE_URL, echo=False)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    
    async with async_session() as db:
        # Verify Navas exists and get correct name
        navas_res = await db.execute(select(Fornitore).where(Fornitore.id == 7))
        navas = navas_res.scalar_one_or_none()
        supplier_name = navas.nome_azienda if navas else "Navas Distribuzione srl"
        
        # Query for all matched products of Navas
        stmt = (
            select(
                RigaFattura.sku_interno,
                func.max(RigaFattura.descrizione_fornitore_raw).label("descrizione"),
                func.min(RigaFattura.prezzo_netto_normalizzato).label("prezzo_min"),
                func.max(RigaFattura.prezzo_netto_normalizzato).label("prezzo_max"),
                # Weighted average: SUM(prezzo * quantita) / SUM(quantita)
                func.sum(RigaFattura.prezzo_netto_normalizzato * RigaFattura.quantita).label("totale_speso"),
                func.sum(RigaFattura.quantita).label("quantita_totale"),
                func.count(RigaFattura.id).label("numero_acquisti"),
                func.max(Fattura.data_documento).label("ultimo_acquisto")
            )
            .join(Fattura, RigaFattura.fattura_id == Fattura.id)
            .where(
                and_(
                    Fattura.fornitore_id == 7,
                    RigaFattura.stato_matching == StatoMatching.matched,
                    RigaFattura.sku_interno.isnot(None),
                    RigaFattura.prezzo_netto_normalizzato > 0,
                    RigaFattura.is_omaggio.isnot(True)
                )
            )
            .group_by(RigaFattura.sku_interno)
            .order_by(RigaFattura.sku_interno)
        )
        
        res = await db.execute(stmt)
        rows = res.all()
        
        # Load active contract prices from ListinoMaster for Navas
        listino_stmt = select(ListinoMaster).where(
            and_(
                ListinoMaster.fornitore_id == 7,
                ListinoMaster.data_scadenza.is_(None)
            )
        )
        listino_res = await db.execute(listino_stmt)
        listini = listino_res.scalars().all()
        contract_prices = {l.sku_interno: l for l in listini}
        
        # Excel Creation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Storico Prezzi Navas"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"REPORT STORICO PREZZI - {supplier_name.upper()}"
        title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Subtitle Info block
        ws["A2"] = "Fornitore:"
        ws["A2"].font = Font(name="Segoe UI", bold=True)
        ws["B2"] = f"{supplier_name} (ID: 7)"
        
        ws["A3"] = "Data Generazione:"
        ws["A3"].font = Font(name="Segoe UI", bold=True)
        ws["B3"] = date.today().strftime("%d/%m/%Y")
        
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        
        # Headers (Row 5)
        headers = [
            "SKU Interno",
            "Descrizione Prodotto",
            "Prezzo Contr. Attivo",
            "Prezzo Min",
            "Prezzo Max",
            "Prezzo Medio (Semplice)",
            "Prezzo Medio (Pesato)",
            "Volume Acquistato",
            "Fatturato Totale",
            "Ultimo Acquisto"
        ]
        
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Corporate Blue
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
        ws.row_dimensions[5].height = 28
        
        # Fill Data starting from row 6
        current_row = 6
        
        fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Zebra light slate
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Text alignments
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        for item in rows:
            sku = item.sku_interno
            desc = item.descrizione
            p_min = float(item.prezzo_min)
            p_max = float(item.prezzo_max)
            tot_speso = float(item.totale_speso)
            qta_tot = float(item.quantita_totale)
            n_acq = item.numero_acquisti
            ultimo = item.ultimo_acquisto.strftime("%d/%m/%Y") if isinstance(item.ultimo_acquisto, date) else str(item.ultimo_acquisto)
            
            # Weighted average
            p_avg_weighted = tot_speso / qta_tot if qta_tot > 0 else 0.0
            
            # Simple average requires querying all prices for this SKU and averaging them
            avg_simple_res = await db.execute(
                select(func.avg(RigaFattura.prezzo_netto_normalizzato))
                .join(Fattura, RigaFattura.fattura_id == Fattura.id)
                .where(
                    and_(
                        Fattura.fornitore_id == 7,
                        RigaFattura.sku_interno == sku,
                        RigaFattura.stato_matching == StatoMatching.matched,
                        RigaFattura.prezzo_netto_normalizzato > 0,
                        RigaFattura.is_omaggio.isnot(True)
                    )
                )
            )
            p_avg_simple = float(avg_simple_res.scalar() or 0.0)
            
            # Active contract price
            listino_item = contract_prices.get(sku)
            p_contract = float(listino_item.prezzo_pattuito) if listino_item else None
            
            # Write to cells
            r_fill = fill_even if current_row % 2 == 0 else fill_odd
            
            c_sku = ws.cell(row=current_row, column=1, value=sku)
            c_desc = ws.cell(row=current_row, column=2, value=desc)
            
            c_contract = ws.cell(row=current_row, column=3)
            if p_contract is not None:
                c_contract.value = p_contract
                c_contract.number_format = '€ #,##0.00'
            else:
                c_contract.value = "—"
                
            c_min = ws.cell(row=current_row, column=4, value=p_min)
            c_min.number_format = '€ #,##0.00'
            
            c_max = ws.cell(row=current_row, column=5, value=p_max)
            c_max.number_format = '€ #,##0.00'
            
            c_avg_simple = ws.cell(row=current_row, column=6, value=p_avg_simple)
            c_avg_simple.number_format = '€ #,##0.00'
            
            c_avg_weighted = ws.cell(row=current_row, column=7, value=p_avg_weighted)
            c_avg_weighted.number_format = '€ #,##0.00'
            
            c_qta = ws.cell(row=current_row, column=8, value=qta_tot)
            c_qta.number_format = '#,##0.0'
            
            c_tot = ws.cell(row=current_row, column=9, value=tot_speso)
            c_tot.number_format = '€ #,##0.00'
            
            c_last = ws.cell(row=current_row, column=10, value=ultimo)
            
            # Apply styling and borders
            for col_num in range(1, 11):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = Font(name="Segoe UI", size=9.5)
                cell.fill = r_fill
                cell.border = thin_border
                
                if col_num in (1, 10):
                    cell.alignment = align_center
                elif col_num == 2:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
        # Add Summary Row at the end
        ws.cell(row=current_row, column=1, value="TOTALE COMPLESSIVO").font = Font(name="Segoe UI", size=10, bold=True)
        ws.cell(row=current_row, column=1).alignment = align_left
        
        # Sum total quantity and total spend formulas
        c_tot_qty = ws.cell(row=current_row, column=8, value=f"=SUM(H6:H{current_row-1})")
        c_tot_qty.number_format = '#,##0.0'
        c_tot_qty.font = Font(name="Segoe UI", size=10, bold=True)
        c_tot_qty.alignment = align_right
        
        c_tot_spent = ws.cell(row=current_row, column=9, value=f"=SUM(I6:I{current_row-1})")
        c_tot_spent.number_format = '€ #,##0.00'
        c_tot_spent.font = Font(name="Segoe UI", size=10, bold=True)
        c_tot_spent.alignment = align_right
        
        summary_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )
        
        for col_num in range(1, 11):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = summary_border
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.row_dimensions[current_row].height = 24
        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Check length of values
            for cell in col:
                val = cell.value
                if val:
                    if str(val).startswith("="):
                        val_len = 15 # estimate size for formulas
                    else:
                        val_len = len(str(val))
                    if val_len > max_len:
                        max_len = val_len
            
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Specific overrides for widths
        ws.column_dimensions["A"].width = 22 # SKU
        ws.column_dimensions["B"].width = 45 # Description
        ws.column_dimensions["C"].width = 16 # Contract Price
        ws.column_dimensions["J"].width = 15 # Last Purchase
        
        # Save Workbook
        wb.save(OUTPUT_PATH)
        print(f"Excel report successfully generated and saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    asyncio.run(main())
