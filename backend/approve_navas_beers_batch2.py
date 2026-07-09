import asyncio
import argparse
from decimal import Decimal
from datetime import date
from datetime import datetime
import openpyxl
from sqlalchemy import select, func

from app.database import async_session_factory
from app.models.products import Product, SupplierProductAlias, MatchCandidate
from app.models.listino import ListinoMaster
from app.services.supplier_list_import import save_append_only_price

BEER_BATCH_2_CONFIG = [
    {
        "supplier_code": "NAVAS_BIR0013",
        "sku": "BEER-HEINEKEN_LAGER-33CL-BT",
        "canonical_name": "Heineken Lager 33 cl bottiglia",
        "brand": "Heineken",
        "category": "beer",
        "volume_ml": 330,
        "pack_qty": 24,
        "container_type": "glass_bottle",
        "comparison_unit": "liter",
        "price": Decimal("17.21"),
        "raw_desc": "BIRRA HEINEKEN 0.33 CL X 24 BT",
        "normalized_desc": "birra heineken 0.33 centilitro 24 pezzi bottiglia",
        "orig_vol": 3,
        "orig_pack": 24,
        "override_reason": "Prezzo/volume normalizzato corretto a 330ml causa refuso listino 0.33 CL"
    },
    {
        "supplier_code": "NAVAS_BIR0017",
        "sku": "BEER-NASTRO_AZZURRO-33CL-BT",
        "canonical_name": "Nastro Azzurro 33 cl bottiglia",
        "brand": "Nastro Azzurro",
        "category": "beer",
        "volume_ml": 330,
        "pack_qty": 24,
        "container_type": "glass_bottle",
        "comparison_unit": "liter",
        "price": Decimal("15.57"),
        "raw_desc": "BIRRA NASTRO AZZ. 0.33 CL X 24 BT",
        "normalized_desc": "birra nastro azz. 0.33 centilitro 24 pezzi bottiglia",
        "orig_vol": 3,
        "orig_pack": 24,
        "override_reason": "Prezzo/volume normalizzato corretto a 330ml causa refuso listino 0.33 CL"
    },
    {
        "supplier_code": "NAVAS_BIR0020",
        "sku": "BEER-PERONI-33CL-BT",
        "canonical_name": "Peroni 33 cl bottiglia",
        "brand": "Peroni",
        "category": "beer",
        "volume_ml": 330,
        "pack_qty": 24,
        "container_type": "glass_bottle",
        "comparison_unit": "liter",
        "price": Decimal("13.52"),
        "raw_desc": "BIRRA PERONI  033X 24 BT",
        "normalized_desc": "birra peroni 033x 24 bottiglia",
        "orig_vol": None,
        "orig_pack": 33,
        "override_reason": "Volume impostato a 330ml e pack corretto a 24 causa errore parsing '033X'"
    },
    {
        "supplier_code": "NAVAS_BIRRA",
        "sku": "BEER-TENNENTS_SUPER-33CL-BT",
        "canonical_name": "Tennent's Super 33 cl bottiglia",
        "brand": "Tennent's",
        "category": "beer",
        "volume_ml": 330,
        "pack_qty": 24,
        "container_type": "glass_bottle",
        "comparison_unit": "liter",
        "price": Decimal("28.69"),
        "raw_desc": "BIRRA TENNENT'S 33 CL",
        "normalized_desc": "birra tennent's 33 centilitro",
        "orig_vol": 330,
        "orig_pack": 1,
        "override_reason": "Prezzo pack/fardello impostato a 24 anziché 1 causa prezzo rilevato per cartone intero"
    }
]

async def check_uniqueness_navas_birra(db):
    stmt = select(SupplierProductAlias).where(
        SupplierProductAlias.supplier_id == 11,
        SupplierProductAlias.supplier_code == "NAVAS_BIRRA"
    )
    res = await db.execute(stmt)
    aliases = res.scalars().all()
    
    excel_count = 0
    try:
        wb = openpyxl.load_workbook("data/import_samples/storico_prezzi_navas.xlsx", data_only=True)
        sheet = wb.active
        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r, column=1).value
            if val == "NAVAS_BIRRA":
                excel_count += 1
    except Exception as e:
        print(f"Errore lettura Excel: {e}")
        
    return len(aliases), excel_count

async def run_batch(apply: bool):
    supplier_id = 11  # Navas Srl
    today = date.today()
    
    print("=" * 100)
    print(f"🍺 NAVAS BEER BATCH 2 - {'APPLY' if apply else 'DRY RUN'} (MANUAL OVERRIDES)")
    print("=" * 100)
    
    async with async_session_factory() as db:
        async with db.begin():
            # 1. Verifica univocità codice NAVAS_BIRRA per Tennent's
            db_aliases, excel_rows = await check_uniqueness_navas_birra(db)
            print(f"🔍 Verifica codice 'NAVAS_BIRRA' per Tennent's:")
            print(f"  - Alias trovati a DB: {db_aliases}")
            print(f"  - Righe trovate in Excel: {excel_rows}")
            
            tennents_valid = True
            if db_aliases > 0 or excel_rows > 1:
                print("  ⚠️ WARNING: Il codice 'NAVAS_BIRRA' non appare univoco! Tennent's andrà in 'manual_supplier_code_review'.")
                tennents_valid = False
            else:
                print("  ✅ Il codice 'NAVAS_BIRRA' è univoco.")
            print("-" * 100)
            
            # Conteggi prima
            prod_count_before = (await db.execute(select(func.count(Product.id)))).scalar()
            alias_count_before = (await db.execute(select(func.count(SupplierProductAlias.id)).where(SupplierProductAlias.supplier_id == supplier_id))).scalar()
            price_count_before = (await db.execute(select(func.count(ListinoMaster.id)).where(ListinoMaster.fornitore_id == supplier_id, ListinoMaster.data_scadenza.is_(None)))).scalar()
            alias_no_price_before = (await db.execute(
                select(func.count(SupplierProductAlias.id))
                .outerjoin(ListinoMaster, ListinoMaster.supplier_product_alias_id == SupplierProductAlias.id)
                .where(SupplierProductAlias.supplier_id == supplier_id, SupplierProductAlias.status == "approved", ListinoMaster.id.is_(None))
            )).scalar()
                
            print("\n📊 CONTEGGI DB PRE-OPERAZIONE:")
            print(f"  Prodotti Canonici Totali: {prod_count_before}")
            print(f"  Alias Navas Totali:        {alias_count_before}")
            print(f"  Prezzi Navas Attivi:       {price_count_before}")
            print(f"  Alias Approved Senza Prezzo: {alias_no_price_before}")
            print("-" * 100)
            
            products_created = 0
            aliases_created = 0
            prices_created = 0
            
            for conf in BEER_BATCH_2_CONFIG:
                if conf["supplier_code"] == "NAVAS_BIRRA" and not tennents_valid:
                    print(f"⏭️ SALTO TENNENT'S: Codice fornitore non univoco.")
                    continue
                    
                print(f"Elaborazione: '{conf['raw_desc']}'...")
                
                # A. Crea/Verifica Prodotto Canonico
                p_stmt = select(Product).where(Product.sku_interno == conf["sku"])
                prod = (await db.execute(p_stmt)).scalars().first()
                
                if not prod:
                    if apply:
                        prod = Product(
                            sku_interno=conf["sku"],
                            canonical_name=conf["canonical_name"],
                            normalized_name=conf["canonical_name"].lower(),
                            brand=conf["brand"],
                            category=conf["category"],
                            volume_ml=conf["volume_ml"],
                            unit_count=1,
                            container_type=conf["container_type"],
                            comparison_unit=conf["comparison_unit"],
                            is_active=True
                        )
                        db.add(prod)
                        await db.flush()
                        print(f"  ➕ [PRODOTTO] Creato: {conf['sku']} -> {conf['canonical_name']}")
                    else:
                        print(f"  🔍 [PRODOTTO] Da Creare: {conf['sku']} -> {conf['canonical_name']}")
                    products_created += 1
                else:
                    print(f"  ✓ [PRODOTTO] Già Esistente: {conf['sku']}")
                
                # B. Crea/Verifica SupplierProductAlias
                a_stmt = select(SupplierProductAlias).where(
                    SupplierProductAlias.supplier_id == supplier_id,
                    SupplierProductAlias.supplier_code == conf["supplier_code"]
                )
                alias = (await db.execute(a_stmt)).scalars().first()
                
                if not alias:
                    if apply:
                        alias = SupplierProductAlias(
                            supplier_id=supplier_id,
                            product_id=prod.id,
                            supplier_code=conf["supplier_code"],
                            raw_description=conf["raw_desc"],
                            normalized_description=conf["normalized_desc"],
                            pack_qty=conf["pack_qty"],
                            volume_ml=conf["volume_ml"],
                            status="approved",
                            confidence_score=100.0,
                            source="manual_override_beers"
                        )
                        db.add(alias)
                        await db.flush()
                        print(f"  ➕ [ALIAS] Creato: {conf['supplier_code']} (pack: {conf['pack_qty']}, vol: {conf['volume_ml']}ml, override: {conf['override_reason']})")
                    else:
                        print(f"  🔍 [ALIAS] Da Creare: {conf['supplier_code']} (pack: {conf['pack_qty']}, vol: {conf['volume_ml']}ml)")
                    aliases_created += 1
                else:
                    print(f"  ✓ [ALIAS] Già Esistente: {conf['supplier_code']}")
                    if apply and alias.product_id != prod.id:
                        alias.product_id = prod.id
                        alias.status = "approved"
                        alias.pack_qty = conf["pack_qty"]
                        alias.volume_ml = conf["volume_ml"]
                        db.add(alias)
                        print(f"  ⚡ [ALIAS] Ricollegato ed aggiornato per override")
                
                # C. Salva Prezzo in ListinoMaster
                vol_litri = Decimal(str(conf["volume_ml"])) / Decimal("1000")
                prezzo_norm = conf["price"] / (Decimal(str(conf["pack_qty"])) * vol_litri)
                
                if apply:
                    outcome = await save_append_only_price(
                        db=db,
                        fornitore_id=supplier_id,
                        sku_interno=conf["sku"],
                        descrizione=conf["raw_desc"],
                        prezzo_pattuito=conf["price"],
                        unita_misura="piece",
                        data_inizio=today,
                        supplier_product_alias_id=alias.id
                    )
                    print(f"  ➕ [PREZZO] Esito: {outcome} | Prezzo Pack: € {conf['price']:.4f} | Normalizzato: € {prezzo_norm:.4f} / L")
                    if outcome in ("created", "updated"):
                        prices_created += 1
                else:
                    print(f"  🔍 [PREZZO] Da Creare: Prezzo Pack = € {conf['price']:.4f} | Normalizzato = € {prezzo_norm:.4f} / L")
                    prices_created += 1
                    
                # D. Risolve MatchCandidate
                mc_stmt = select(MatchCandidate).where(
                    MatchCandidate.supplier_id == supplier_id,
                    MatchCandidate.raw_description == conf["raw_desc"],
                    MatchCandidate.status == "pending"
                )
                mc = (await db.execute(mc_stmt)).scalars().first()
                if mc:
                    if apply:
                        mc.status = "resolved"
                        mc.resolved_at = datetime.utcnow()
                        db.add(mc)
                        print(f"  ⚡ [CANDIDATO] Risolto MatchCandidate per '{conf['raw_desc']}'")
                    else:
                        print(f"  🔍 [CANDIDATO] Verrà risolto MatchCandidate per '{conf['raw_desc']}'")
                print("-" * 50)
                
            if not apply:
                print("\n⚠️ MODALITÀ DRY RUN: Nessun dato salvato (ROLLBACK).")
                # Forza il rollback sollevando un'eccezione silenziata o lasciando che il context manager faccia rollback automatico a fine transazione
            else:
                print("\n✅ MODALITÀ APPLY: Dati salvati con successo (COMMIT).")
                
            # Calcola conteggi post transazione
            if not apply:
                prod_count_after = prod_count_before + products_created
                alias_count_after = alias_count_before + aliases_created
                price_count_after = price_count_before + prices_created
                alias_no_price_after = alias_no_price_before
            else:
                # Nel caso di apply reale, facciamo flush prima di contare
                await db.flush()
                prod_count_after = (await db.execute(select(func.count(Product.id)))).scalar()
                alias_count_after = (await db.execute(select(func.count(SupplierProductAlias.id)).where(SupplierProductAlias.supplier_id == supplier_id))).scalar()
                price_count_after = (await db.execute(select(func.count(ListinoMaster.id)).where(ListinoMaster.fornitore_id == supplier_id, ListinoMaster.data_scadenza.is_(None)))).scalar()
                alias_no_price_after = (await db.execute(
                    select(func.count(SupplierProductAlias.id))
                    .outerjoin(ListinoMaster, ListinoMaster.supplier_product_alias_id == SupplierProductAlias.id)
                    .where(SupplierProductAlias.supplier_id == supplier_id, SupplierProductAlias.status == "approved", ListinoMaster.id.is_(None))
                )).scalar()
                
            print("\n📊 CONTEGGI DB POST-OPERAZIONE:")
            print(f"  Prodotti Canonici Totali: {prod_count_after} (Variazione: +{prod_count_after - prod_count_before})")
            print(f"  Alias Navas Totali:        {alias_count_after} (Variazione: +{alias_count_after - alias_count_before})")
            print(f"  Prezzi Navas Attivi:       {price_count_after} (Variazione: +{price_count_after - price_count_before})")
            print(f"  Alias Approved Senza Prezzo: {alias_no_price_after}")
            print("=" * 100)
            
            # Se siamo in Dry Run, solleviamo un'eccezione fittizia per far scattare il rollback
            if not apply:
                raise RuntimeError("Dry run rollback trigger")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Approve Navas beers batch 2 with overrides")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true", help="Simulate import process")
    group.add_argument("--apply", action="store_true", help="Execute database changes")
    
    args = parser.parse_args()
    try:
        asyncio.run(run_batch(apply=args.apply))
    except RuntimeError as e:
        if str(e) != "Dry run rollback trigger":
            raise
